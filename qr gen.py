'''def random_string(text): import random as ra import itertools as it

import pprint as pprint

 

-perm= list(map("".join, it.permutations(text)))

return ra.sample(perm, len(perm))

 

 

text=raw_input("Enter your text here:")

samp=random_string(text)

print samp '''

 

print "Starting function...importing modules"

 

from openpyxl import load_workbook

import os

from os import path
from os import listdir
from os.path import isfile, join

import pyqrcode as qr
from zipfile import *

import sys

import datetime
import re

 

 

 

 

def alert(value):

    import Tkinter as tk

    #from Tkinter import tkMessageBox

    import tkMessageBox as tkBox

   

    box=tk.Tk()

    if(value =='error'):

        tkBox.showwarning("Program is exiting","There was an issue,Please restart the app")

    if(value.startswith('save')):

        msg="{} values recorded ".format(value[4:])   

        tkBox.showwarning("Save Succesfull",msg)
    else:
        tkBox.showwarning("Save Succesfull",value)
        

    box.destroy()

    return

   

'.........................................'#end of alert function

 

def Qr(serial_arr):

    global home
    print "Generating QR code"


    value = str(datetime.datetime.now())

    val=datetime.datetime.strptime(value[:value.index('.')], "%Y-%m-%d %H:%M:%S").date()

           

    dest=str(val)#os.path.join(home,str(val))

    if os.path.exists(dest):
        #dest= dest+' 2'

        os.chdir(os.path.join(home,dest))

    elif not os.path.exists(dest):

        os.mkdir(dest)
        os.chdir(os.path.join(home,dest))
        alert('new folder created')

    
    for i in range(1,len(serial_arr)):
        serial_str=str(serial_arr[i])+'.png'
        q=qr.create(serial_str)
        q.png(serial_str,scale=8)

 
    
    print '{} QR generated in the folder..{}'.format(len(serial_arr)-1,os.path.join(home,dest))

    os.chdir(home)

    alert('save {} QR created'.format(len(serial_arr)-1))
    return

 

'.........................................'#end of Qr function'

 

def inc(text):

    return str(int(int(text)+1)).zfill(2)

    #return text[:2]+str(int(int(text[2:4])+1)).zfill(2)

 

 

'.........................................'#end of inc function

 

def in_list(serial):

    global excel_global
    #print excel_global
    return any(serial in x for x in excel_global)

 

'.........................................'#end of in_list function

 

def vendor(n):

    arr=[]

    
    strng="IN01"         

    try:
        text=raw_input("Enter 1 character of Vendor Code:").upper()

        if(len(text)!=1 or len(re.findall('\d+', text))>0 or(  text.isalpha() ==False ) ):
            
            return vendor(n)

        
        arr.append('Vendor code: '+text)


        for i in range(0,int(n)):


            serial=strng+text+"KOL"+str(i).zfill(6)

            if in_list(serial):

               print "Serial number '{}' already used .stopping operation".format(serial)

               return arr

            arr.append(serial )

        return arr

    except:

        print "Something went wrong,please enter Vendor code again"

        return vendor(n)

 

       

'.........................................'#end of vendor function

 

def digits(n):

    arr=[]

    digit=raw_input("Enter the last 6 digits:")

    if(len(digit) !=6):

            print "Please exact numbers(including 00xxxx)"

            return digits(n)

    try:
        if( ( digit.isdigit() ==False) or  (int(digit)<0 or int(digit)>999999 )):
            
            print "Please enter only 6 numbers(including 00xxxx)"

            return digits(n)

       

        arr.append('Digits: '+digit)

        strng="NKOL"

        batch= "01"
        digit=abs(int(digit))
        j=0
      

        for i in range(digit,digit+int(n)):
                                
            if(i>=999999 ):
                if(i==999999 ):
                     batch=inc(batch)
                     #strng="IN"+inc(batch)+strng
                     #strng=strng[:2]+str(int(int(strng[2:4])+1)).zfill(2)+strng[4:]
                     '''n=int(n)-i
                     digit=0
                     print n,digit'''
                     
                i=-1
                j=j+1
                i=i+j-1

            temp="IN"+batch+strng+str(i+1).zfill(6)

            if in_list(temp):

                 print "Serial number '{}' already used .stopping operation".format(temp)

                 return arr

 

            arr.append(temp)

        return arr

    except:

        print "Something went wrong,please enter last 6 digits again"

        return digits(n)

 

'.........................................'#end of digits function

 

 

def both(n):

    arr=[]

    text=raw_input("Enter Vendor Code:").upper()

    if(len(text)!=1 or len(re.findall('\d+', text))>0 or (  text.isalpha() ==False) ):

            print "Please exact Vendor code"

            return both(n)

           

    digit=raw_input("Enter the last 6 digits:")

    try:
        if(len(digit) !=6 ):

            print "Please exact numbers(including 00xxxx)"

            return both(n)
         
        if( ( digit.isdigit() ==False) or (int(digit)<0 or  int(digit)>999999 )): #999990
            print "Please enter only 6 numbers(including 00xxxx)"

            return both(n)

        

        arr.append('Vendor Code: '+text+' & Digits: '+digit)

        batch="01"
        digit= int(digit)
        j=0

        for i in range(int(digit),int(digit)+int(n)):
            
            if(i>=999999 ):
                if(i==999999 ):                    
                    batch=inc(batch)
                    #strng="IN"+inc(batch)+strng
                    #strng=strng[:2]+str(int(int(strng[2:4])+1)).zfill(2)+strng[4:]
                    '''n=int(n)-i
                    digit=0
                    print n,digit'''

                i=-1
                j=j+1
                i=i+j-1

                
            temp="IN"+batch+text+"KOL"+str(i+1).zfill(6)
            

            if in_list(temp):

                print "Serial number '{}' already used .stopping operation".format(temp)

                return arr

            arr.append(temp)

        return arr

    except:

        print "Something went wrong,please enter again"

        return both(n)

 

'.........................................'#end of both function

 

 

def full(n):

    arr=[]

    text=raw_input("Enter Full text:").upper()  #IN01NKOL444446

    try:

       

        if(len(text)!=14 or ( text[4:8].isalpha() ==False) ):

            print "Please enter exact text(including 00xxxx)",1

            return full(n)

        arr.append('Full Text: '+text)

      
        

        if(text[2:4].isdigit() and (int(text[2:4])<0 or int(text[2:4])>99 )):

            print "Please enter exact text(including 00xxxx)",2

            return full(n)
 

        digit=text[-6:]
        

        if(( digit.isdigit() ==False) and (len(digit)<6 or len(digit)>6) and (int(digit)<0 or int(digit)>999999 )):

            print "Please enter only 6 numbers(including 00xxxx)"

            return full(n)

        j=0 

        for i in range(int(digit),int(digit)+int(n)):

            if(i>=999999 ):
                if (i==999999):
                    text=text[:2]+inc(text[2:4])+text[4:]
                    #n=int(n)-i
                    #digit=0

                i=-1
                j=j+1
                i=i+j-1

            
            temp=text[:8]+str(int(i+1)).zfill(6)

            if in_list(temp):

                print "Serial number '{}' already used .Stopping operation".format(temp)

                return arr

           

            arr.append(temp)

        return arr

    except:

        print "Please exact text(including 00xxxx)"

        return full(n)

       

'.........................................'#end of full function

 

 

def last(n):

    arr=[]

   

    try:

        global excel_global

        text=excel_global[-1]

        arr.append('Print {} from last run '.format(n))

       

        digit=text[-6:]

        j=0     

        for i in range(int(digit),int(digit)+int(n)):
            if(i>=999999 ):
                if (i==999999):
                    text=text[:2]+inc(text[2:4])+text[4:]
                    #n=int(n)-i
                    #digit=0

                i=-1
                j=j+1
                i=i+j-1

            temp=text[:8]+str(i+1).zfill(6)

            if in_list(temp):

                print "Serial number '{}' already used .stopping operation".format(temp)

                return arr

           

            arr.append(temp)

        return arr

    except:

        print "Unknown error restarting program,run program again if the issue persists"

        return last(n)

 

 

 

'.........................................'#end of last function

 

 

def save_excel(data):

    #import datetime

    #from openpyxl import load_workbook
    global home
    filenam= os.path.join(home,'Serial codes.xlsx')
    #print filenam

    wb = load_workbook(filename = filenam)

    sheet=wb.active

    mx_row=sheet.max_row

 

    try:

       

        if isinstance(data, str):

           

            cont_str=[]

            for row in sheet['E2:F{}'.format(mx_row)]:#quantity,start serial
                for cell in row:
                    

                    temp=str(cell.value).encode('ascii','ignore')
                    

                    if (cell.col_idx ==5):

                        cols_range=int(temp)

                    elif(cell.col_idx ==6):

                        digit=int(temp[-6:])

                        i=0
                        

                        for qty in range(digit,digit+cols_range):

                            temp=temp

                            if ( qty ==1000000 ):

                                temp=temp[:2]+inc(temp[2:4])+temp[4:8]+str(qty).zfill(6)

                                i=0
                            xl_str=temp[:-6]+str(int(int(temp[-6:])+i)).zfill(6)
                            

                            cont_str.append(xl_str)

                            i=i+1

            return cont_str

        if isinstance(data, list):
 
           

            value = str(datetime.datetime.now()).replace('-','/')

            val=datetime.datetime.strptime(value[:value.index('.')], "%Y/%m/%d %H:%M:%S")

            print (len(data),data)

           

            for i in range(len(data)):

                #print ("Hi",mx_row+1,i+1,cell.value,data[i])

                sheet.cell(row=mx_row+1, column=i+1).value = data[i]

                wb.save("Serial codes.xlsx")

            alert('save '+str(data[4]))
            return 

 

    except:

        print "Unknown error restarting program,run program again if the issue persists"

 

 

'.........................................'#end of save_excel function

 

 

def main():

    import sys

    import datetime

    try:

        n=raw_input("Type quantity: \n")

        if(n.isdigit()==False):

            print "Enter only  numbers"

            main()

        print "Do you have: 1. vendor code only or 2.do you have last 6 digits only or 3.both Vendor and last 6 digits. or 4. Do you have the full text \n 5.To print from the last used serial"

        print "Or 'Enter' to type quantity again\n"

 

        inp= raw_input("Please enter the choice number accodingly ")

        if(inp=='1'):

            switcher=vendor(n)

        elif(inp=='2'):

            switcher=digits(n)

        elif(inp=='3'):

            switcher=both(n)

        elif(inp=='4'):

            switcher=full(n)

        elif(inp=='5'):

            switcher=last(n)

        else:

            switcher=main()

        print switcher

 

        logs= raw_input("Do you want to save logs?").lower()

        if (logs.startswith('y')):

           

            name= raw_input("Enter your name: ")

            value = str(datetime.datetime.now()).replace('-','/')

           

            val=datetime.datetime.strptime(value[:value.index('.')], "%Y/%m/%d %H:%M:%S")

            data=[]

            data.append(val)

            data.append(name)

            data.append(inp)

            data.append(switcher[0])

            data.append(n)

            data.append(switcher[1])

            data.append(switcher[-1])

           

            #data.extend(,,,,switcher[1],switcher[-1])

            print "Saving Excel"

            save_excel(data)


            Qr(switcher)

       

        if(raw_input("Press y to reload program and OK to save").startswith('y')):

            main()

    except (KeyboardInterrupt, SystemExit):

        print "You have pressed Ctrl+C,this interupts the program, please restart application again"

 

        alert('error')

 

        sys.exit()

 

 

    except:
        print "Starting program again"

        main()

 

'.........................................'#end of main function

 
def upgrade():
    
    
    ready= raw_input("Do you want to check for updates?")
    if(ready.startswith('y')):
        internet()
    else:
        return
 

def internet():
    import requests
    global home
    #files = [f for f in listdir(home) if isfile(join(home, f)) and f.endswith(".txt")]
    files= open("code_versions.txt","a+")
    f=files.read().splitlines()[2:]
    vers_code=[v.split(':')[1] for v in f]
    files.close()    
    
    #os.path.join(home,'version.txt')
    try:
        url="https://spreadsheets.google.com/feeds/list/15d4q_oluuKxbyF6texIPeIHW3kI0-urO81mn902d89o/od6/public/values?alt=json"
        

        response=requests.get(url, allow_redirects=True).json()
        
        d=response['feed']
        
        g_rows=d['openSearch$totalResults']['$t']
        entry=d['entry']
        

        rels= []
        fileurl=[]
        filename=[]
        vers = []
        prev=[]
        Data = {}
        for row in entry:
            ct=0
            for key in row:
                if (ct==0):
                    
                    rels.append(str(row['gsx$released']['$t']))
                    filename.append(str(row['gsx$filename']['$t']))
                    vers.append(str(row['gsx$versionname']['$t']))
                    prev.append(str(row['gsx$previousversion']['$t']))
                    fileurl.append(str(row['gsx$fileurl']['$t']))
                ct=ct+1

        Data.update({'fileurl': fileurl,'vers': vers,'prev': prev,'name':filename,'rels':rels})
       
        ver_upd=set(vers_code).issubset(Data['name'])


        if(len(vers_code)== len(Data['vers']) or  ver_upd):
            print "Scripts already running latest scripts"
            return
        else :
            
            items= Data['name'][int(g_rows)-1]
            print "in else",items
            
            url2="https://drive.google.com/uc?export=download&id="+Data['fileurl'][int(g_rows)-1]
            dwn=requests.get(url2, allow_redirects=True)
            alert("Downloading latest file")
                                        
            
            open(items, 'wb').write(dwn.content)
                                        
            zip_name=Data['vers'][int(g_rows)-1]+".zip"
            zip_ar=ZipFile(zip_name,"w")  
            zip_ar.write(items)
            zip_ar.close()
            
            
            #url='https://drive.google.com/uc?export=download&id=1IcYwzUknkrdjVVbIMaPR804aFKDRh85J'
            new_line=str(Data['rels'][int(g_rows)-1])+":"+str(Data['vers'][int(g_rows)-1])+"\r\n"
            print new_line
            files1= open("code_versions.txt","a+")
            files1.write(new_line)
            print("Version updated!!")
            files1.close()
            if os.path.isfile(items):
                os.remove(items)
            else:
                print ("no such files to delete")


    except OSError as e:  ## if failed, report it back to the user ##
        print ("Error: %s - %s." % (e.filename, e.strerror))
                           
    except:
        return "Not able to connect to internet"    
    

    

#global variables

home =os.path.abspath(os.path.dirname(sys.argv[0]))#sys.path[0] #path.realpath("Serial codes.xlsx")

excel_global= save_excel('true')

print "All modules imported, starting..."
upgrade()

main()

